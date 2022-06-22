from win32com.client import Dispatch
from openpyxl import load_workbook
from PIL import ImageGrab
from openpyxl.utils import get_column_letter, column_index_from_string
import geopandas as gpd
import pandas as pd
import numpy as np
import numbers
from datetime import datetime
from IPython.display import display
from tqdm import tqdm
import sys
import csv

script_folder = r'C:\General\BCC_Software\Python\python_repository\python_library\python_main_bcc'
if script_folder not in sys.path: sys.path.append(script_folder)
from project_tools.intersection_turn_counts import gis_tools as gis
from project_tools.intersection_turn_counts import file_utls as fu

# https://docs.microsoft.com/en-us/office/vba/api/office.msoshapetype
shape_type_dict = {30: 'mso3DModel',
                   1: 'msoAutoShape',
                   2: 'msoCallout',
                   20: 'msoCanvas',
                   3: 'msoChart',
                   4: 'msoComment',
                   27: 'msoContentApp',
                   21: 'msoDiagram',
                   7: 'msoEmbeddedOLEObject',
                   8: 'msoFormControl',
                   5: 'msoFreeform',
                   28: 'msoGraphic',
                   6: 'msoGroup',
                   24: 'msoIgxGraphic',
                   22: 'msoInk',
                   23: 'msoInkComment',
                   9: 'msoLine',
                   31: 'msoLinked3DModel',
                   29: 'msoLinkedGraphic',
                   10: 'msoLinkedOLEObject',
                   11: 'msoLinkedPicture',
                   16: 'msoMedia',
                   12: 'msoOLEControlObject',
                   13: 'msoPicture',
                   14: 'msoPlaceholder',
                   18: 'msoScriptAnchor',
                   -2: 'msoShapeTypeMixed',
                   25: 'msoSlicer',
                   19: 'msoTable',
                   17: 'msoTextBox',
                   15: 'msoTextEffect',
                   26: 'msoWebVideo'}

arrow_head_style = {1: 'msoArrowheadNone',
                    2: 'msoArrowheadTriangle',
                    3: 'msoArrowheadOpen',
                    4: 'msoArrowheadStealth',
                    5: 'msoArrowheadDiamond',
                    6: 'msoArrowheadOval',
                    -2: 'msoArrowheadStyleMixed'}

# https://docs.microsoft.com/en-us/office/vba/api/excel.xlhalign
excel_horizontal_alignment_dict = {-4108: 'Center',
                                   7: 'Center across selection',
                                   -4117: 'Distribute',
                                   5: 'Fill',
                                   1: 'Align according to data type',
                                   -4130: 'Justify',
                                   -4131: 'Left',
                                   -4152: 'Right'}

excel_border_checks = {'xlDiagonalDown': 5, 'xlDiagonalUp': 6, 'xlEdgeBottom': 9, 'xlEdgeLeft': 7, 'xlEdgeRight': 10,
                       'xlEdgeTop': 8, 'xlInsideHorizontal': 12, 'xlInsideVertical': 11}


def ole_to_date_str(pywin_datetime):
    py_datetime = datetime(
        year=pywin_datetime.year,
        month=pywin_datetime.month,
        day=pywin_datetime.day,
        hour=pywin_datetime.hour,
        minute=pywin_datetime.minute,
        second=pywin_datetime.second)
    date_str = py_datetime.strftime("%d/%m/%Y")
    return date_str


def find_number_of_columns(ws, row_no, col_no):
    while ws.cells(row_no, col_no).value is None:
        col_no += 1
    return col_no


def check_border(ws, row, col, location=9):
    border_style = ws.cells(row, col).Borders(location).LineStyle
    return border_style


def check_cell_value_strings(ws, row_from, search_strings, loop_rows=None, col_no=1):
    search_strings_lower = [str(x).lower() for x in search_strings]
    row_no_ = row_from
    output_row = -1
    if loop_rows is None:
        loop_rows = 100
    while row_no_ <= row_from + loop_rows and output_row < 0:

        cell_value = ws.cells(row_no_, col_no).value
        if cell_value is None:
            if search_strings is list:
                if None in search_strings:
                    output_row = row_no_
        elif isinstance(search_strings, list):
            # any(ext in url_string for ext in extensionsToCheck)
            if any(str(string_item).lower() in str(cell_value).lower() for string_item in search_strings_lower):
                output_row = row_no_
        else:
            if search_strings.lower() in str(cell_value).lower():
                output_row = row_no_
        row_no_ += 1
    return output_row


def create_header(ws, header_top, rows, col_left, end_col):
    # ToDo:  May cause errors for peds if 3 teirs of data
    rows = [*range(header_top, header_top + rows + 1, 1)]
    cols = [*range(col_left, end_col + 1, 1)]
    header_1 = []
    header_2 = []
    previous_head = None
    head_dict_for_row = {}
    # border_style = check_border(ws, row_no_, col_no, location)
    for c in cols:
        top_row = True
        head = ""
        for r in rows:
            if top_row:
                cell_value = ws.cells(r, c).value
                if cell_value is not None:
                    header_1.append(str(cell_value))
                    previous_head = cell_value
                    top_row = False
                elif previous_head is not None:
                    header_1.append(previous_head)
                    top_row = False
            else:
                if ws.cells(r, c).Mergecells:
                    cell_value = ws.cells(r, c).MergeArea.Cells(1, 1).Value
                else:
                    cell_value = ws.cells(r, c).value
                if cell_value is not None:
                    if cell_value not in head and cell_value not in header_1:
                        if head == '':
                            head = cell_value
                        else:
                            head += f'_{cell_value}'
        head.replace('None|', '').replace('|None', '')
        header_2.append(head)
    head_array = [header_1, header_2]
    return head_array


def find_end_row_from_border(ws, row_no_, col_no=1, location=8, next_row=1):
    border_style = -4142
    additional_rows = 0
    while border_style == -4142 and additional_rows < 10:
        border_style = check_border(ws, row_no_, col_no, location)  # 8 is location for top of cell refer top doc
        if border_style == -4142:
            additional_rows += next_row
            row_no_ = row_no_ + next_row
    return row_no_


def find_location_of_data_in_spreadsheet(ws, row_no, header_strings, df_end_strings, col_no=1, survey_type=None):
    search_for_additional_data = True
    header_end = None
    end_df_row = None
    header = None
    use_columns = None
    header_row = check_cell_value_strings(ws, row_no, search_strings=header_strings, loop_rows=40, col_no=col_no)
    if header_row == -1:
        search_for_additional_data = False
    else:
        header_row = find_end_row_from_border(ws, header_row, col_no=col_no, location=8, next_row=-1)

        row_no = header_row
        header_end = check_cell_value_strings(ws, row_no, search_strings=['Time Period', '(1/4 hr end)'], loop_rows=40,
                                              col_no=col_no)
        # https://docs.microsoft.com/en-us/dotnet/api/microsoft.office.interop.excel.xllinestyle?view=excel-pia

        use_columns = find_columns_used(ws, header_row, col_from=col_no)
        if survey_type == 'matrix_1':
            header_row += 1  # ignore street name
        # display(use_columns)
        start_col = column_index_from_string(use_columns.split(':')[0])
        end_col = column_index_from_string(use_columns.split(':')[1])
        header = create_header(ws, header_row, header_end - header_row, start_col, end_col)
        # header = create_header(ws, header_row, header_end - header_row, 1, header_end)
        end_df_row = check_cell_value_strings(ws, row_no, search_strings=df_end_strings, loop_rows=None,
                                              col_no=col_no) - 1
    return row_no, search_for_additional_data, end_df_row, header, header_end, use_columns


def create_dataframe_ausraffic(excel_file_path, sheet_name, header_end, end_df_row, use_columns, header):
    data_df = pd.read_excel(excel_file_path,
                            sheet_name=sheet_name,
                            skiprows=header_end,
                            header=None,
                            index_col=None,
                            nrows=end_df_row - header_end,
                            usecols=use_columns,
                            parse_dates=False)
    if data_df is None:
        return None
    else:

        data_df.columns = pd.MultiIndex.from_arrays(header)
        data_df.columns = data_df.columns.map(lambda x: '|'.join([str(i) for i in x]))
        # display(1, data_df)
        data_df = data_df.reset_index(drop=True)
        if data_df.columns[0].lower() == 'time|(1/4 hr end)':
            data_df = data_df.rename(columns={data_df.columns[0]: 'TIME|(1/4 hr end)'})
        df_melt = data_df.melt(id_vars=['TIME|(1/4 hr end)'], var_name='spreadsheet_movement|vehicle',
                               value_name='count')
        # display(2, df_melt)
        df_melt[['temp_spreadsheet_movement', 'temp_vehicle']] = df_melt[
            'spreadsheet_movement|vehicle'].str.split('|',
                                                      expand=True)

        df_melt['vehicle'] = np.where(
            df_melt['temp_spreadsheet_movement'].str.lower().str.contains('pedestrian'),
            'pedestrian', df_melt['temp_vehicle'])
        df_melt['spreadsheet_movement'] = np.where(
            df_melt['temp_spreadsheet_movement'].str.lower().str.contains('pedestrian'),
            df_melt['temp_vehicle'], df_melt['temp_spreadsheet_movement'])

        return df_melt


def create_dataframe_matrix(excel_file_path, sheet_name, header_end, end_df_row, use_columns, header):
    data_df = pd.read_excel(excel_file_path,
                            sheet_name=sheet_name,
                            skiprows=header_end,
                            header=None,
                            index_col=None,
                            nrows=end_df_row - header_end,
                            usecols=use_columns,
                            parse_dates=False)
    headers_test = pd.MultiIndex.from_arrays(header)
    if data_df is None:
        return None
    else:
        header_items = len(header[0])
        columns_in_df = len(data_df.columns)
        if header_items > columns_in_df:
            items_to_remove = header_items - columns_in_df
            top_header = header[0][:-items_to_remove]
            bottom_header = header[1][:-items_to_remove]
            header = [top_header, bottom_header]
        data_df.columns = pd.MultiIndex.from_arrays(header)
        data_df.columns = data_df.columns.map(lambda x: '|'.join([str(i) for i in x]))
        # display(1, data_df)
        data_df = data_df.reset_index(drop=True)
        col_1 = data_df.columns[0].lower()
        col_2 = data_df.columns[1].lower()
        col_3 = data_df.columns[2].lower()
        if col_1 == col_2 == col_3 == 'direction|time period':
            data_df = data_df.iloc[:, ~data_df.columns.duplicated()]
            data_df = data_df.rename(columns={data_df.columns[0]: 'TIME|(1/4 hr end)'})
        df_melt = data_df.melt(id_vars=['TIME|(1/4 hr end)'], var_name='spreadsheet_movement|vehicle',
                               value_name='count')
        # display(2, df_melt)
        df_melt[['temp_spreadsheet_movement', 'temp_vehicle']] = df_melt[
            'spreadsheet_movement|vehicle'].str.split('|',
                                                      expand=True)

        df_melt['vehicle'] = np.where(
            df_melt['temp_spreadsheet_movement'].str.lower().str.contains('pedestrian'),
            'pedestrian', df_melt['temp_vehicle'])
        df_melt['spreadsheet_movement'] = np.where(
            df_melt['temp_spreadsheet_movement'].str.lower().str.contains('pedestrian'),
            df_melt['temp_vehicle'], df_melt['temp_spreadsheet_movement'])

        return df_melt


def find_count_data_austraffic(excel_file_path, sheet_name, ws, row_from, header_strings, df_end_strings, col_no=1):
    search_for_additional_data = True
    spreadsheet_df = None
    dataframes = []
    row_no = row_from
    while search_for_additional_data:
        spreadsheet_location_info = find_location_of_data_in_spreadsheet(ws, row_no, header_strings, df_end_strings,
                                                                         col_no=col_no)
        search_for_additional_data = spreadsheet_location_info[1]

        if search_for_additional_data:
            row_no = spreadsheet_location_info[0]
            end_df_row = spreadsheet_location_info[2]
            header = spreadsheet_location_info[3]
            header_end = spreadsheet_location_info[4]
            use_columns = spreadsheet_location_info[5]
            row_no = end_df_row
            data_df = pd.read_excel(excel_file_path, sheet_name=sheet_name, skiprows=header_end, header=None,
                                    index_col=None, nrows=end_df_row - header_end, usecols=use_columns,
                                    parse_dates=False)
            # pd.MultiIndex.from_arrays(header)
            if data_df is None:
                return None
            else:
                df_melt = create_dataframe_ausraffic(excel_file_path, sheet_name, header_end, end_df_row, use_columns,
                                                     header)
                dataframes.append(df_melt)
    if dataframes:
        spreadsheet_df = pd.concat(dataframes)
    return spreadsheet_df


def get_survey_info_matrix(ws):
    survey_site = ws.cells(6, 5).value.replace(':', '').strip()
    survey_date = ws.cells(8, 5).value.replace(':', '').strip()
    survey_weather = ws.cells(9, 5).value.replace(':', '').strip()
    if survey_site[:1].isnumeric():
        survey_site = survey_site.split(' ', 1)[1]
    survey_info_dict = {'survey_weather': survey_weather, 'survey_date': survey_date, 'survey_site': survey_site}

    return survey_info_dict


def get_survey_info_austraffic(ws):
    row_loop = range(1, 11)
    col_loop = range(1, 11)
    survey_info_dict = {'survey_site': None, 'survey_date': None, 'survey_weather': None, 'lat': None, 'lon': None,
                        'geocode_location': None}
    for col in col_loop:
        for row in row_loop:
            cell_value = ws.cells(row, col).value
            if 'location' in str(cell_value).lower():
                survey_info_dict['survey_site'] = ws.cells(row, col + 1).value
            elif 'date' in str(cell_value).lower():
                pywin_datetime = ws.cells(row, col + 1).value
                py_date = ole_to_date_str(pywin_datetime)
                survey_info_dict['survey_date'] = py_date
            elif 'weather' in str(cell_value).lower():
                survey_info_dict['survey_weather'] = ws.cells(row, col + 1).value

    return survey_info_dict


def pandas_read_excel_multi_index_with_use_cols(excel_file, sheet_name=0, header_from=0, header_to=0, n_rows=None,
                                                use_cols=None):
    """
    Pandas is unable to specify usecols when extracting multi-index dataframe from excel.  This function
    provides a workaround for this feature.
    Parameters
    ----------
    excel_file (string): Path to excel file to be read
    sheet_name (string or int): name (string) or index (int) of excel sheet to be read
    header_from (int): first line of haeder.
    header_to (int): last line of header
    n_rows (int): number of rows to be included in dataframe
    use_cols (string or int): rows to be used in dataframe

    Returns
    -------
    object (dataframe)
    """
    # df = pd.read_excel(excel_file, sheet_name=sheet_name, header=[header_from, header_to], nrows=n_rows, usecols=use_cols
    try:
        df = pd.read_excel(excel_file,
                           sheet_name=sheet_name,
                           header=header_to,
                           index_col=[header_from, header_to],
                           nrows=n_rows,
                           usecols=use_cols,
                           parse_dates=False)
        index = pd.read_excel(excel_file,
                              sheet_name=sheet_name,
                              header=None,
                              skiprows=header_from,
                              index_col=[header_from, header_to],
                              nrows=2,
                              usecols=use_cols,
                              parse_dates=False)

        index = index.fillna(method='ffill', axis=1)
        df.columns = pd.MultiIndex.from_arrays(index.values)
        df.columns = df.columns.map(lambda x: '|'.join([str(i) for i in x]))
        df = df.reset_index(drop=True)
    except:
        df = None
    return df


def find_columns_used_ss(ws, top_row, bottom_row, col_from=1):
    # ToDo: Delete this function if the function below works for all cases.  Otherwise use this for Austraffic.
    col_no = col_from
    # if ws.cells(row_no, col_no).value is None:
    #    row_no += 1
    col_end_found = False
    while not col_end_found:
        row_no = top_row
        valid_cells = False
        while row_no <= bottom_row and not valid_cells:
            if ws.cells(row_no, col_no).value is not None:
                col_no += 1
                valid_cells = True

            else:
                if row_no == bottom_row:
                    col_end_found = True
                row_no += 1
    start_col = get_column_letter(col_from)
    end_col = get_column_letter(col_no - 1)

    use_columns = f'{start_col}:{end_col}'
    return use_columns


def find_columns_used(ws, top_row, col_from=1):
    # if ws.cells(row_no, col_no).value is None:
    #    row_no += 1
    col_no = col_from
    # if ws.cells(row_no, col_no).value is None:
    #    row_no += 1
    border_location = excel_border_checks['xlEdgeTop']
    border_style = None
    col_end_found = False
    while border_style != -4142:
        border_style = check_border(ws, top_row, col_no, border_location)  # 8 is location for top of cell refer top doc
        if border_style != -4142:
            col_no += 1
    start_col = get_column_letter(col_from)
    end_col = get_column_letter(col_no - 1)

    use_columns = f'{start_col}:{end_col}'
    return use_columns


'''
def convert_spreadsheet_to_dataframe(excel_file_path, ws, sheet_name=0, search_strings_headings='Time',
                                     search_strings_end=[None, 'Peak', 'Total'], row_no=1, loop_rows=30):
    # ToDo: check if this is used

    while row_no != -1:
        row_details = find_count_data_rows(excel_file_path, sheet_name, ws, row_from=1, loop_rows=30,
                                           header_strings=search_strings_headings, df_end_strings=search_strings_end)

        data_df = pandas_read_excel_multi_index_with_use_cols(excel_file_path, sheet_name, header_from=row_no - 1,
                                                              header_to=row_no, n_rows=5, use_cols=use_columns)
        row_no = row_details[-1]
'''


def find_col_widths(ws, start_range=1, end_range=31):
    col_positions = []
    current_width = 0
    for col in range(start_range, end_range):
        cell_width = ws.cells(1, col).Width
        col_positions.append(current_width)
        current_width += cell_width
    return col_positions


def find_nearest_position(locations, position):
    if type(locations) is dict:
        locations = list(locations.keys())
    locations = np.array(locations)
    position = np.array(position)
    distances = np.linalg.norm(locations - position, axis=1)
    min_index = np.argmin(distances)
    closest_point = locations[min_index]
    closest_distance = distances[min_index]
    closest_point = (closest_point[0], closest_point[1])
    return closest_point, closest_distance


def find_row_positions(ws, start_range=1, end_range=31):
    row_positions = []
    row_hieght = 0
    for row in range(start_range, end_range):
        cell_height = ws.cells(row, 1).Height
        row_positions.append(row_hieght)
        row_hieght += cell_height
    return row_positions


def find_cell_from_positions(row_pos, cell_pos, row_positions, col_positions):
    print('dod')


def find_nearest(list_, value):
    if type(list_) is dict:
        locations = list(list_.keys())
    array = np.asarray(list_)
    if array.max() < value:
        cell_location = None
    else:
        min_value_greater = array[array > value].min()
        cell_location = list_.index(min_value_greater)
    return cell_location


def get_text_box_locations_in_dictionary(ws, shapes):
    text_box_dict = {}
    col_positions = find_col_widths(ws, start_range=1, end_range=31)
    # print(col_positions)
    row_positions = find_row_positions(ws, start_range=1, end_range=31)
    # print(row_positions)
    for sh in shapes:
        if shape_type_dict[sh.Type] == 'msoTextBox':
            shape_top = sh.top
            shape_height = sh.height
            shape_left = sh.left
            shape_width = sh.width
            # print('\n', 'text box checks: ', sh.name, shape_top, )
            shape_pos_row = shape_top + shape_height / 2
            shape_pos_col = shape_left + shape_width / 2
            pos_row = find_nearest(row_positions, shape_pos_row)
            pos_col = find_nearest(col_positions, shape_pos_col)
            text_box_value = str(ws.TextBoxes(sh.Name).Text)
            # print('text box checks: ', sh.name, )
            # print('text box checks shape_top: ', shape_top)
            # print('text box checks shape_height: ', shape_height)
            # print('text box checks shape_height: ', shape_pos_row)

            if pos_row is not None and pos_col is not None and text_box_value.replace('.', '', 1).isdigit():
                text_box_dict[(shape_pos_row, shape_pos_col)] = text_box_value
    if text_box_dict == {}:
        text_box_dict = None
    return text_box_dict


def ungroup_shapes(ws, shapes):
    for sh in shapes:
        if shape_type_dict[sh.Type] == 'msoGroup':
            sh.ungroup()
    shapes_2 = ws.Shapes
    return shapes_2


def create_movement_dict(ws):
    movement_dict = {}
    shapes_1 = ws.shapes
    shapes = ungroup_shapes(ws, shapes_1)
    text_box_dict = get_text_box_locations_in_dictionary(ws, shapes)
    # display(text_box_dict)

    # shapes = ws.shapes
    for sh in shapes:
        if shape_type_dict[sh.Type] == 'msoLine':
            line_name = sh.Name
            end_style = sh.Line.EndArrowheadStyle
            begin_style = sh.Line.BeginArrowheadStyle
            cell_col = sh.TopLeftCell.Address.split("$")[1]
            cell_row = sh.TopLeftCell.Address.split("$")[2]
            if (arrow_head_style[end_style] == 'msoArrowheadTriangle' and
                arrow_head_style[begin_style] == 'msoArrowheadNone') and int(cell_row) <= 31:
                # print('begin_style', begin_style)

                cell = (int(cell_row), int(column_index_from_string(cell_col)))
                # print('cell: ' , sh.TopLeftCell.ColumnWidth)
                shape_top = sh.top
                shape_height = sh.height
                shape_left = sh.left
                shape_width = sh.width
                hor_flip = sh.HorizontalFlip
                ver_flip = sh.VerticalFlip
                line_pos = find_point_positions(hor_flip, ver_flip, shape_top, shape_top + shape_height, shape_left,
                                                shape_left + shape_width)
                # print('line details: ', line_name, line_pos)
                point_1_for_angle_calc = (
                    (-line_pos[0][0], line_pos[0][1]))  # flip vertical to get positive distance in north direction
                point_2_for_angle_calc = ((-line_pos[1][0], line_pos[1][1]))
                angle = gis.compass_angle(point_1_for_angle_calc, point_2_for_angle_calc, excel_cell_format=True)
                angle_round = int(gis.custom_round(angle, base=45))
                direction_dict = gis.get_direction_dict()
                direction = direction_dict[angle_round]
                #print('arrow info: ', sh.name)
                movement, approach = find_turn_movement(ws, cell, arrow_direction=direction, arrow_points=line_pos,
                                                        text_dict=text_box_dict)
                # correct direction for slip turn arrows
                if direction in ['NE', 'SE', 'SW', 'NW']:
                    if approach == 'N':
                        if direction == 'SE':
                            direction = 'E'
                        elif direction == 'SW':
                            direction = 'W'
                    elif approach == 'E':
                        if direction == 'NW':
                            direction = 'N'
                        elif direction == 'SW':
                            direction = 'S'
                    elif approach == 'S':
                        if direction == 'NE':
                            direction = 'E'
                        elif direction == 'NW':
                            direction = 'W'
                    elif approach == 'W':
                        if direction == 'NE':
                            direction = 'N'
                        elif direction == 'SE':
                            direction = 'S'
                movement_dict[str(movement)] = f"{approach}_{direction}"
    return movement_dict


def add_qld_aus_to_geolocate_text(geo_search_text):
    if 'qld' not in geo_search_text.lower() or 'queensland' not in geo_search_text.lower():
        geo_search_text += f'{geo_search_text}, QLD'
    if 'aus' not in geo_search_text.lower() or 'australia' not in geo_search_text.lower():
        geo_search_text += ', AUSTRALIA'
    return geo_search_text


def add_geocode(geo_search_text):
    lat, lon, location = gis.geocode_coordinates(geo_search_text, user_agent='Engineering_Services_BCC',
                                                 api='google')
    if location is not None:
        location = location[0]
    return lat, lon, location


def add_site_info(df, survey_info_dict, movement_dict):
    df['spreadsheet_movement'] = df['spreadsheet_movement'].str.replace("movement ", "", case=False)
    df['movement'] = df['spreadsheet_movement'].map(movement_dict)
    df['intersection'] = survey_info_dict['survey_site']
    df['date'] = survey_info_dict['survey_date']
    df['weather'] = survey_info_dict['survey_weather']
    df['intersection_lat'] = survey_info_dict['lat']
    df['intersection_lon'] = survey_info_dict['lon']
    # df_melt['geolocated_location'] = location

    gdf = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df.intersection_lon, df.intersection_lat), crs='epsg:4326')
    # df = pd.DataFrame.from_dict([movement_dict])
    return gdf


def get_austraffic_1_survey_data(excel_file_path, sheet_name):
    # ToDo: updated doc string
    """
    find relationship between turn movement numbers and origin to destination approach.  Approach is designated by North
    (N), East (E), South (S) or West (W).
    Parameters
    ----------
    excel_file_path (string): string of path to excel file

    Returns
    -------
    DataFrame: {movement_1: Origin_Destination, movement_2, origin_destination}
    """
    xl = Dispatch('Excel.Application')
    #xl.Interactive = False
    #xl.Visible = False
    wb = xl.Workbooks.Open(Filename=excel_file_path)
    ws = wb.Worksheets(sheet_name)
    movement_dict = create_movement_dict(ws)
    df = find_count_data_austraffic(excel_file_path, sheet_name, ws, row_from=1, header_strings='Time',
                                    df_end_strings=[None, 'Peak', 'Total'])
    if df is None:
        wb.Close(True)
        df_melt = None
        movement_dict = None
        survey_info_dict = None
    else:
        survey_info_dict = get_survey_info_austraffic(ws)
        # intersection = survey_info_dict['survey_site']
        df_melt = df.drop(columns=['temp_spreadsheet_movement', 'temp_vehicle', 'spreadsheet_movement|vehicle'],
                          axis=1)

        df_melt = df_melt.rename(columns={'TIME|(1/4 hr end)': 'survey_time'})
        df_melt['spreadsheet_movement'] = df_melt['spreadsheet_movement'].str.replace("movement ", "", case=False)
        df_melt['file_name'] = excel_file_path
        df_melt['sheet_name'] = sheet_name
        wb.Close(True)
        #xl.Interactive = True
        #xl.Visible = True

    return df_melt, movement_dict, survey_info_dict


def get_ttm_1_survey_data(excel_file_path, sheet_name):
    df = pd.DataFrame
    return df


def find_count_data_matrix(excel_file_path, sheet_name, ws, row_from, header_strings, df_end_strings, col_no=2):
    search_for_additional_data = True
    spreadsheet_df = None
    dataframes = []
    row_no = row_from
    header_strings = 'approach'
    df_end_strings = 'total'
    while search_for_additional_data:
        spreadsheet_location_info = find_location_of_data_in_spreadsheet(ws, row_no, header_strings, df_end_strings,
                                                                         col_no=col_no, survey_type='matrix_1')
        search_for_additional_data = spreadsheet_location_info[1]

        if search_for_additional_data:
            row_no = spreadsheet_location_info[0]
            end_df_row = spreadsheet_location_info[2]
            header = spreadsheet_location_info[3]
            header_end = spreadsheet_location_info[4]
            use_columns = spreadsheet_location_info[5]
            row_no = end_df_row
            data_df = pd.read_excel(excel_file_path, sheet_name=sheet_name, skiprows=header_end, header=None,
                                    index_col=None, nrows=end_df_row - header_end, usecols=use_columns,
                                    parse_dates=False)
            # pd.MultiIndex.from_arrays(header)
            if data_df is None:
                return None
            else:
                df_melt = create_dataframe_matrix(excel_file_path, sheet_name, header_end, end_df_row, use_columns,
                                                  header)
                dataframes.append(df_melt)
    if dataframes:
        spreadsheet_df = pd.concat(dataframes)
    return spreadsheet_df


def get_matrix_1_survey_data(excel_file_path, sheet_name):
    movement_dict = {'1': 'S_W',
                     '2': 'S_N',
                     '3': 'S_E',
                     '3U': 'S_S',
                     '4': 'E_W',
                     '5': 'E_N',
                     '6': 'E_E',
                     '6U': 'E_S',
                     '7': 'N_W',
                     '8': 'N_N',
                     '9': 'N_E',
                     '9U': 'N_S',
                     '10': 'W_W',
                     '11': 'W_N',
                     '12': 'W_E',
                     '12U': 'W_S'
                     }
    xl = Dispatch('Excel.Application')
    # xl.Interactive = False
    # xl.Visible = False
    wb = xl.Workbooks.Open(Filename=excel_file_path)
    ws = wb.Worksheets(sheet_name)
    df = find_count_data_matrix(excel_file_path, sheet_name, ws, row_from=1, header_strings='Approach',
                                df_end_strings=[None, 'Peak', 'Total'])
    if df is None:
        wb.Close(True)
        movement_dict = None
        survey_info_dict = None
    else:
        survey_info_dict = get_survey_info_matrix(ws)
    df = df[df['count'].notna()]
    df = df.drop(columns=['temp_spreadsheet_movement', 'temp_vehicle', 'spreadsheet_movement|vehicle'],
                 axis=1)

    df = df.rename(columns={'TIME|(1/4 hr end)': 'survey_time'})
    df.loc[:, 'spreadsheet_movement'] = df['spreadsheet_movement'].str.replace("direction ", "", case=False)
    df.loc[:, 'spreadsheet_movement'] = df['spreadsheet_movement'].str.split(r"\r").str[0]
    df.loc[:, 'spreadsheet_movement'] = df['spreadsheet_movement'].str.split(r"\n").str[0]
    df.loc[:, 'spreadsheet_movement'] = np.where(df['vehicle'].str.contains(' to '), df['vehicle'], df['spreadsheet_movement'])
    df.loc[:, 'vehicle'] = np.where(df['vehicle'].str.contains(' to '), 'peds', df['vehicle'])
    # df['spreadsheet_movement'] = df['spreadsheet_movement'].str.replace("direction ", "", case=False).split("\r", 1)[0]
    df['file_name'] = excel_file_path
    df['sheet_name'] = sheet_name

    wb.Close(True)
    return df, movement_dict, survey_info_dict


def get_data_audit_systems_1_survey_data(ws):
    df = pd.DataFrame
    return df


survey_functions_map = {"austraffic_1": get_austraffic_1_survey_data, "ttm_1": get_ttm_1_survey_data,
                        "matrix_1": get_matrix_1_survey_data,
                        " data_audit_systems_1": get_data_audit_systems_1_survey_data}


def get_sheets_in_workbook(file_path):
    if not file_path.endswith('.xls'):
        wb = load_workbook(filename=file_path)
        all_sheets = wb.sheetnames
        sheets = []
        for sheet in all_sheets:
            if wb[sheet].sheet_state == 'visible':
                sheets.append(sheet)
    else:
        sheets = None
    return sheets


def get_survey_data_main(excel_file_path, sheet):
    # ToDo: add capability for .xls files
    df = None
    movement_dict = None
    site_info_dict = None
    if 'summary' not in sheet.lower():
        if not excel_file_path.endswith('.xls'):
            survey_format = find_survey_type(excel_file_path, sheet)
            if survey_format is not None:
                df, movement_dict, site_info_dict = survey_functions_map[survey_format](excel_file_path,
                                                                                        sheet_name=sheet)
                if movement_dict is not None:
                    if 'None' in movement_dict:
                        # ToDo: update movement dict for text boxes??
                        # df = None
                        movement_dict = None
            else:
                df = None
                movement_dict = None

    return df, movement_dict, site_info_dict


def find_survey_type(excel_file_path, sheet):
    wb = load_workbook(filename=excel_file_path)
    survey_format = None

    try:
        a1_value = wb[sheet]["A1"].value
    except:
        # ToDo: update this.  this was a quick fix for chart sheet error
        print(f'update this.  This was a quick fix for chart sheet error for sheet {sheet}.')
        return survey_format

    if a1_value is not None:

        if str(a1_value).lower() == 'austraffic video intersection count':
            survey_format = 'austraffic_1'

    else:
        e10_value = wb[sheet]["E10"].value
        e11_value = wb[sheet]["E11"].value
        if e10_value == ': Classified Intersection Count' and e11_value == ': 15 mins Data':
            survey_format = 'matrix_1'
    wb.close()
    return survey_format


def variable_is_number(no):
    """
    Determines if a string is a number

    Parameters
    ----------
    no (string): any variable of type string

    Returns
    -------
    Bool: True if variable is number, False if it is not a number.
    """
    return isinstance(no, numbers.Number)


def check_text_boxes_for_movement():
    print('check_text_boxes')


def approach_from_text_box_position(from_point, to_point, split_compass_45=False):
    angle = gis.compass_angle(from_point, to_point, excel_cell_format=True)
    if split_compass_45:
        angle_round = int(gis.custom_round(angle, base=45))
    else:
        angle_round = int(gis.custom_round(angle, base=90))
    direction_dict = gis.get_direction_dict()
    direction = direction_dict[angle_round]
    return direction


def use_bearing_to_determine_best_match_from_2_nearest_points(point_1, point_2, start_point):
    angle_point_1 = gis.compass_angle(start_point, point_1)
    angle_point_1_round = int(gis.custom_round(angle_point_1, base=90))
    angle_point_2 = gis.compass_angle(start_point, point_2)
    angle_point_2_round = int(gis.custom_round(angle_point_2, base=90))
    angle_dif_point_1 = min(abs(angle_point_1 - angle_point_1_round), abs(angle_point_1 - angle_point_1_round + 360))
    angle_dif_point_2 = min(abs(angle_point_2 - angle_point_2_round), abs(angle_point_2 - angle_point_2_round + 360))
    if angle_dif_point_1 <= angle_dif_point_2:
        return point_1
    else:
        return point_2


def find_turn_movement(worksheet, cell, arrow_direction=None, arrow_points=None, text_dict=None):
    """
    Finds the turn movement number and approach for a turn arrow object within an intersection turning survey
    Parameters
    ----------
    worksheet (excel worksheet object): worksheet object being assessed
    cell (row, col): cell value in which the arrow belongs

    Returns
    -------
    Turn movement number associated to the arrow and approach the turn belongs to
    """
    movement = None
    approach = None
    cell_value = worksheet.cells(cell[0], cell[1]).value
    if cell_value is not None and not isinstance(cell_value, str):  # this is to capture through movements where the arrow starts in the movement cell
        movement = int(cell_value)
        if arrow_direction == 'N':
            approach = 'S'
        elif arrow_direction == 'E':
            approach = 'W'
        elif arrow_direction == 'S':
            approach = 'N'
        elif arrow_direction == 'W':
            approach = 'E'
        else:
            approach = None
    else:
        e_movement = worksheet.cells(cell[0], cell[1] + 1).value
        w_movement = worksheet.cells(cell[0], cell[1] - 1).value
        if cell[0] > 1:
            n_movement = worksheet.cells(cell[0] - 1, cell[1]).value
        else:
            n_movement = None
        if n_movement is None:
            if cell[0] > 2:
                n_movement = worksheet.cells(cell[0] - 2, cell[1]).value
        s_movement = worksheet.cells(cell[0] + 2, cell[1]).value
        if s_movement is None:
            s_movement = worksheet.cells(cell[0] + 1, cell[1]).value

        if variable_is_number(n_movement):
            movement = int(n_movement)
            approach = 'N'
        elif variable_is_number(e_movement):
            movement = int(e_movement)
            approach = 'E'
        elif variable_is_number(s_movement):
            movement = int(s_movement)
            approach = 'S'
        elif variable_is_number(w_movement):
            movement = int(w_movement)
            approach = 'W'
    if movement is None:
        if arrow_points is not None and text_dict is not None:
            rows = [arrow_points[0][0], arrow_points[1][0]]
            cols = [arrow_points[0][1], arrow_points[1][1]]
            # print('arrow_direction:', arrow_direction)
            if arrow_direction == 'N':
                if rows[0] > rows[1]:
                    start_point = arrow_points[0]
                else:
                    start_point = arrow_points[1]
            elif arrow_direction == 'S':
                if rows[0] > rows[1]:
                    start_point = arrow_points[1]
                else:
                    start_point = arrow_points[0]
            elif arrow_direction == 'E':
                if cols[0] < cols[1]:
                    start_point = arrow_points[0]
                else:
                    start_point = arrow_points[1]
            elif arrow_direction == 'W':
                if cols[0] < rows[1]:
                    start_point = arrow_points[1]
                else:
                    start_point = arrow_points[0]
            else:
                start_point = arrow_points[0]  # this will be used for ne, se, sw or nw arrows
            text_positions = list(text_dict.keys())
            nearest_point = find_nearest_position(text_positions, start_point)
            nearest = nearest_point[0]
            nearest_distance = nearest_point[1]
            if nearest_distance <= 15:  # arbitrary number 15 seems to workf or distance between
                point_match = nearest
            elif len(text_dict) >= 2:
                text_dict_2 = text_dict.copy()
                del text_dict_2[nearest]
                text_positions_2 = list(text_dict_2.keys())
                second_nearest = find_nearest_position(text_positions_2, start_point)[0]
                point_match = use_bearing_to_determine_best_match_from_2_nearest_points(nearest, second_nearest,
                                                                                        start_point)

            else:
                point_match = nearest
            movement = text_dict[point_match]
            point_1_for_angle_calc = (
                -start_point[0], start_point[1])  # flip vertical to get positive distance in north direction
            point_2_for_angle_calc = (-point_match[0], point_match[1])
            approach = approach_from_text_box_position(point_1_for_angle_calc, point_2_for_angle_calc)
    return movement, approach


def find_point_positions(hor, ver, top, bottom, left, right):
    """
    Find the start and end positions of an 'msoline' from MS excel
    Parameters
    ----------
    hor: represents the horizontal lip of the line.  0 if left to right, -1 if right to left.
    ver: represents the horizontal lip of the line.  0 if top to bottom, -1 if bottom to top.
    top: y position of the top point.
    bottom: y position of the bottom point
    left: x position of the left most point
    right: x position of the right most point

    Returns
    -------

    """
    point_1 = (top, left)
    point_2 = (top, right)
    point_3 = (bottom, left)
    point_4 = (bottom, right)
    hor_vert_str = f"{hor}_{ver}"
    point_dict = {'0_0': [point_1, point_4],
                  '0_-1': [point_3, point_2],
                  '-1_0': [point_2, point_3],
                  '-1_-1': [point_4, point_1]}
    line_pos = point_dict[hor_vert_str]
    return line_pos


def get_excel_image(file_path, sheet_name, range=None):
    xl = Dispatch('Excel.Application')
    wb = xl.Workbooks.Open(file_path)
    ws = wb.Worksheets[sheet_name]
    ws.Range(ws.Cells(1, 8), ws.Cells(15, 36)).Copy()
    img = ImageGrab.grabclipboard()
    wb.Application.CutCopyMode = False
    wb.Close(True)
    return img


def check_approach_match(xl_approaches, gis_approaches_4, gis_approaches_8=None):
    print('todo')


def check_log_file_exists_create_return_df(file_path, log_type=0):
    """
    Checks if log file exists.  If the file does not exist, a new log file is created.  The func
    Parameters
    ----------
    file_path (string): filepath as string

    Returns
    -------

    """

    path_type = fu.check_file_path_is_folder_or_directory(file_path)
    file_exists = False
    df = None
    if path_type == 'file':
        file_exists = True
    if not file_exists:
        with open(file_path, 'w') as f:
            if log_type == 0 or log_type.lower() == 'movement':
                f.write('"file","sheet","intersection","movement","i","j","k","comments"')
            elif log_type == 1 or log_type.lower() == 'data':
                f.write('"file","sheet","intersection","movement","i","j","k","comments"')
    return


def save_data_log(df, filename):
    with open(filename, 'a', newline='') as f:
        df.to_csv(f, mode='a', header=f.tell() == 0, index=False, encoding="utf-8", quoting=csv.QUOTE_ALL)
    return


def save_movement_log(df, filename):
    with open(filename, 'a', newline='') as f:
        df.to_csv(f, mode='a', header=f.tell() == 0, index=False, encoding="utf-8", quoting=csv.QUOTE_ALL)
    return


def find_files_assessed(df):
    files_assessed = df[df['file'].notnull()]
    return files_assessed


def find_lat_long_from_meta(file_path):
    wb = load_workbook(filename=file_path)
    lat = wb['META']['A2'].value
    lon = wb['META']['B2'].value
    return lat, lon

def exclude_files_already_assessed(all_files, assessed_file, col_check='file_name'):
    df_analysed = pd.read_csv(assessed_file, encoding='cp1252')
    df_file_names = df_analysed[col_check].unique().tolist()
    new_files = list(set(all_files) - set(df_file_names))
    return new_files


def analyse_intersection_counts_for_saturn(file_path, sections_file, nodes_file, log_file_data, log_file_movements,
                                           from_file=None, test_run=False):
    path_type = fu.check_file_path_is_folder_or_directory(file_path)
    if path_type is not None:
        if path_type == 'file':
            files = [file_path]
        elif path_type == 'directory':
            files = fu.get_list_of_files_in_directory(file_path, file_type='.xl*', sub_folders=False)

        if from_file is not None:
            files = files[from_file:]
        if fu.check_file_exists(log_file_movements):
            files = exclude_files_already_assessed(all_files=files, assessed_file=log_file_movements)

        for f in tqdm(files):
            if test_run:
                print(f)
            sheet_names = get_sheets_in_workbook(f)
            geo_location = None

            if sheet_names is not None:
                if 'META' in sheet_names:
                    lat_lon = find_lat_long_from_meta(f)
                    geo_location = 'META'
                else:
                    lat_lon = None
                for sheet in sheet_names:
                    if test_run:
                        print('sheet: ', sheet)
                    lat = None
                    lon = None
                    # print(sheet)
                    if 'summary' not in sheet.lower():
                        movement_ijk_dict = None
                        survey_df, movement_dict, survey_info_dict = get_survey_data_main(f, sheet)
                        # print(movement_dict, survey_info_dict, survey_df)
                        if survey_df is not None:
                            if lat_lon is not None:
                                lat = lat_lon[0]
                                lon = lat_lon[1]
                            else:
                                geo_search_text = add_qld_aus_to_geolocate_text(survey_info_dict['survey_site'])
                                lat, lon, geo_location = add_geocode(geo_search_text)
                            survey_info_dict['lat'] = lat
                            survey_info_dict['lon'] = lon
                            survey_info_dict['geocode_location'] = geo_location
                            #survey_df_pivot = clean_data_output(survey_df)
                            # survey_df_2 = add_site_info(survey_df, survey_info_dict, movement_dict)
                            if not test_run:
                                save_data_log(survey_df, log_file_data)
                        else:
                            survey_info_dict = {'survey_site': None, 'survey_date': None, 'survey_weather': None,
                                                'lat': None, 'lon': None, 'geocode_location': None}
                        if movement_dict is not None:
                            if movement_dict == {}:

                                movement_ijk_dict = {'excel_movement': [None], 'geographic_movement': [None],
                                                     'angle_from': [None], 'angle_to': [None], 'i': [None], 'j': [None],
                                                     'k': [None], 'log_type': [0.1], 'dist_to_node': None}
                                movement_log_df = pd.DataFrame.from_dict(movement_ijk_dict)
                                movement_log_df.loc[:, 'spreadsheet_approach_from_to'] = None

                            else:
                                survey_df_2 = add_site_info(survey_df, survey_info_dict, movement_dict)
                                sections_gdf = gis.create_sections_gdf(sections_file)
                                nodes_gdf = gis.create_nodes_gdf(nodes_file, crs='epsg:4326')
                                sections_gdf = gis.find_node_start_and_end(sections_gdf, nodes_gdf)
                                node_distance_gdf = gis.find_node_distance_from_intersection(nodes_gdf,
                                                                                             survey_df=survey_df_2)
                                movement_ijk_dict = gis.find_ijk(sections_gdf, node_distance_gdf, survey_df,
                                                                 movement_dict=movement_dict)
                                movement_log_df = pd.DataFrame.from_dict(movement_ijk_dict)
                                movement_log_df.loc[:, 'spreadsheet_approach_from_to'] = movement_log_df[
                                    'excel_movement'].map(movement_dict)


                        else:
                            movement_ijk_dict = {'excel_movement': [None], 'geographic_movement': [None],
                                                 'angle_from': [None], 'angle_to': [None], 'i': [None], 'j': [None],
                                                 'k': [None], 'log_type': [0], 'dist_to_node': None}
                            movement_log_df = pd.DataFrame.from_dict(movement_ijk_dict)
                            movement_log_df.loc[:, 'spreadsheet_approach_from_to'] = None
                        movement_log_df.loc[:, 'file_name'] = f
                        movement_log_df.loc[:, 'sheet_name'] = sheet
                        movement_log_df.loc[:, 'intersection'] = survey_info_dict.get('survey_site')
                        movement_log_df.loc[:, 'geocode_location'] = geo_location
                        movement_log_df.loc[:, 'lat'] = lat
                        movement_log_df.loc[:, 'lon'] = lon
                        movement_log_df.loc[:, 'survey_date'] = survey_info_dict.get('survey_date')
                        movement_log_df.loc[:, 'survey_date'] = pd.to_datetime(movement_log_df['survey_date'])
                        movement_log_df.loc[:, 'day'] = np.where(movement_log_df['survey_date'].isnull(), np.nan,
                                                                 movement_log_df['survey_date'].dt.day_name())
                        movement_log_df.loc[:, 'weather'] = survey_info_dict.get('survey_weather')

                        in_proj, out_proj = gis.create_in_out_projections_for_conversion("EPSG:4326", "EPSG:28356")
                        movement_log_df.loc[:, 'map_info_location'] = movement_log_df.apply(
                            lambda row: gis.add_map_info_coords(row['lat'], row['lon'], in_proj, out_proj), axis=1)
                        movement_log_df = movement_log_df[
                            ['intersection', 'excel_movement', 'spreadsheet_approach_from_to', 'geographic_movement',
                             'survey_date', 'day', 'weather', 'geocode_location', 'dist_to_node', 'lat', 'lon', 'angle_from', 'angle_to', 'i',
                             'j', 'k', 'log_type', 'file_name', 'sheet_name', 'map_info_location']]
                        if not test_run:
                            save_movement_log(movement_log_df.sort_values(by=['excel_movement']), log_file_movements)
                        else:
                            display(movement_log_df)
    # ToDo: check movements match data movements


def clean_data_output(df):
    df['text_check'] = df.apply(lambda row: check_is_numeric(str(row['count'])), axis=1)
    df = df.dropna(subset = ['count'])
    df = df[(df['vehicle'].str.lower().str.contains('ped')==False) | (df['vehicle'].str.lower().str.contains('cyclist')==False)]
    df = df[df['text_check'].notnull()]
    df = df.astype({'count': float})
    display(df.head(), df.info())
    df['survey_time'] = df['survey_time'].astype("str")
    df['survey_time'] = df['survey_time'].str.split(' ').str[-1].str.split('.').str[0]
    df_pivot = pd.pivot_table(df, values=['count'], columns='survey_time',
                              index=['vehicle', 'spreadsheet_movement', 'file_name', 'sheet_name'],
                              aggfunc={'count': np.mean})
    df_pivot = flatten_multi_index_columns(df_pivot)
    return df_pivot


def flatten_multi_index_columns(df):
    columns = df.columns.map(lambda x: '|'.join([str(i) for i in x])).tolist()
    new_columns = []
    for c in columns:
        new_columns.append(c.split("|")[-1])

    # df_pivot.reset_index()
    df.columns = new_columns
    df.reset_index()
    return df


def check_is_numeric(text):
    text.strip()
    is_number = True
    decimal_count = 0
    for s in text:
        if not s.isnumeric():
            if s != '.' and decimal_count == 0:
                is_number = False
            else:
                decimal_count += 1

    if is_number:
        return text
    else:
        return None